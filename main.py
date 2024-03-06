import telebot
import random
import string
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

# Установите токен вашего бота
bot = telebot.TeleBot('7199324736:AAG_j6wOSURnCmVMd6wewJ2zcs_e2GR5LYs')

# Путь к файлу Excel
excel_file_path = 'C:/Users/MerV1ng/OneDrive/Рабочий стол/TELEGA.xlsx'

# Токен вашего ЮKassa
yookassa_token = '381764678:TEST:79311'
# ID вашего магазина в ЮKassa
shop_id = '506751'
# ID вашего товара в ЮKassa
shop_article_id = '538350'

# Обработчик команды /start
@bot.message_handler(commands=['start'])
def start(message):
    # Текст приветствия
    welcome_text = """Привет!

Для получения доступа к VPN, нажмите на кнопку "Купить" и следуйте подсказкам, в них будет описан процесс подключения к вашему серверу.

О сервисе:
— Никаких ограничений по количеству устройств и трафику
— VPN работает с Wi-Fi и мобильным интернетом
— Ваши данные шифруются и не доступны для просмотра даже администраторам сервиса
— VPN совместим со всеми популярными приложениями
"""
    # Отправляем сообщение с приветствием и клавиатурой
    bot.send_message(message.chat.id, welcome_text, reply_markup=create_menu_keyboard())

# Обработчик команды "Купить"
@bot.message_handler(func=lambda message: message.text == 'Купить')
def buy(message):
    # Создаем инлайн-кнопки с разными тарифами
    inline_keyboard = InlineKeyboardMarkup()
    # Добавляем кнопки с тарифами
    inline_keyboard.row(
        InlineKeyboardButton("Тариф A - $5", callback_data="tariff_A"),
        InlineKeyboardButton("Тариф B - $10", callback_data="tariff_B")
    )
    inline_keyboard.row(
        InlineKeyboardButton("Тариф C - $15", callback_data="tariff_C"),
        InlineKeyboardButton("Тариф D - $20", callback_data="tariff_D")
    )

    # Отправляем фото и кнопки с тарифами
    bot.send_photo(chat_id=message.chat.id, photo=open('./photo_2023-10-30_09-13-28.jpg', 'rb'), caption="Выберите тариф:", reply_markup=inline_keyboard)

# Обработчик нажатия на инлайн-кнопки тарифов
@bot.callback_query_handler(func=lambda call: call.data.startswith('tariff_'))
def process_tariff_callback(callback_query):
    tariff = callback_query.data.split('_')[1]  # Получаем выбранный тариф
    # Получаем имя пользователя
    username = callback_query.from_user.username
    # Формируем ссылку для оплаты через ЮKассу
    ykassa_payment_url = create_ykassa_payment_url(tariff)
    # Отправляем пользователю сообщение с ссылкой на оплату через ЮKассу
    bot.send_message(callback_query.from_user.id, f"Нажмите на кнопку, чтобы оплатить через ЮKassa:", reply_markup=create_ykassa_payment_keyboard(ykassa_payment_url))
    # Сохраняем данные о покупке
    save_payment_data(callback_query.from_user.id, username, tariff)  # Fix: Pass 'tariff' parameter here


# Функция для создания ссылки на оплату через ЮKассу
def create_ykassa_payment_url(tariff):
    payment_description = ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))
    return f'https://oplata.yookassa.ru/payments/checkout?merchant_id={yookassa_token}&amount={get_tariff_amount(tariff) * 100}&currency=RUB&description={payment_description}&shopId={shop_id}&shopArticleId={shop_article_id}&return_url=https://your.return.url'

# Функция для создания клавиатуры с кнопкой оплаты через ЮKassa
def create_ykassa_payment_keyboard(ykassa_payment_url):
    inline_keyboard = InlineKeyboardMarkup()
    inline_keyboard.row(InlineKeyboardButton("Оплатить через ЮKassa", url=ykassa_payment_url))
    return inline_keyboard

# Функция для получения суммы для каждого тарифа
def get_tariff_amount(tariff):
    if tariff == 'A':
        return 5
    elif tariff == 'B':
        return 10
    elif tariff == 'C':
        return 15
    elif tariff == 'D':
        return 20
    else:
        return 0  # В случае неправильного тарифа вернуть 0 или другое значение

# Функция для сохранения данных о пользователе, тарифе и дате оплаты
def save_payment_data(user_id, username, tariff):
    # Проверяем, существует ли уже файл Excel
    try:
        # Если файл существует, пробуем загрузить его
        wb = load_workbook(excel_file_path)
        sheet = wb.active
    except FileNotFoundError:
        # Если файл не существует, создаем новый и добавляем заголовки
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'Username'
        sheet['B1'] = 'User ID'
        sheet['C1'] = 'Тариф'
        sheet['D1'] = 'Дата оплаты'

    # Находим первую свободную строку для записи новых данных
    next_row = 1
    while sheet[f'B{next_row}'].value:  # Пока столбец B в строке next_row не пустой
        next_row += 1

    # Записываем данные о пользователе, тарифе и дате оплаты в файл Excel
    sheet[f'A{next_row}'] = username
    sheet[f'B{next_row}'] = user_id
    sheet[f'C{next_row}'] = tariff
    sheet[f'D{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Сохраняем изменения в файле Excel
    wb.save(excel_file_path)

    # Создаем инлайн-кнопку "ИНСТРУКЦИЯ" с нужной ссылкой
    inline_keyboard = InlineKeyboardMarkup()
    inline_keyboard.add(InlineKeyboardButton("ИНСТРУКЦИЯ", url="https://teletype.in/@qwertyland/UfTzz5oRBiq"))

    # Отправляем сообщение с текстом и кнопкой "ИНСТРУКЦИЯ"
    bot.send_message(user_id, "Спасибо за покупку! \nЧтобы увидеть свой токен перейдите Мои заказы\nДля настройки подключения к вашему серверу по токену нажмите на кнопку ниже", reply_markup=inline_keyboard)


# Обработчик команды "Полезная информация"
@bot.message_handler(func=lambda message: message.text == 'Полезная информация')
def send_useful_info(message):
    # Создаем клавиатуру с встроенными ссылками
    inline_keyboard = InlineKeyboardMarkup()
    # Добавляем кнопки с встроенными ссылками
    inline_keyboard.row(
        InlineKeyboardButton("Частые вопросы", url="https://teletype.in/@qwertyland/S0SD5G6rTBi"),
        InlineKeyboardButton("Подключение через OUTLINE", url="https://teletype.in/@qwertyland/UfTzz5oRBiq")
    )
    inline_keyboard.row(
        InlineKeyboardButton("Политика конфиденциальности", url="https://teletype.in/@qwertyland/PWJKpKmIsrV")
    )

    # Отправляем сообщение с клавиатурой
    bot.send_message(message.chat.id, "Выберите интересующий раздел:", reply_markup=inline_keyboard)


# Обработчик команды "Мои заказы"
@bot.message_handler(func=lambda message: message.text == 'Мои заказы')
def my_orders(message):
    # Получаем id пользователя
    user_id = message.from_user.id

    # Путь к файлу Excel
    excel_file_path = 'TELEGA.xlsx'

    # Загружаем файл Excel
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    # Создаем список для хранения всех заказов пользователя
    user_orders = []

    # Ищем все заказы пользователя
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=7):
        if str(row[1].value) == str(user_id):
            user_orders.append({
                "Тариф": row[2].value,
                "Дата оплаты": row[3].value,  # Не вызываем strftime() здесь
                "Токен": row[6].value  # Получаем токен из столбца G
            })

    # Если найдены заказы пользователя
    if user_orders:
        # Получаем информацию о последнем заказе пользователя
        latest_order = user_orders[-1]
        order_text = f"Ваш тариф: {latest_order['Тариф']}\nПоследняя дата оплаты: {latest_order['Дата оплаты']}\nВаш токен: {latest_order['Токен']}"

        # Отправляем сообщение с информацией о заказе и токене
        bot.send_message(user_id, order_text)
    else:
        bot.send_message(user_id, "У вас нет заказов.")


# Обработчик команды "Поддержка"
@bot.message_handler(func=lambda message: message.text == 'Поддержка')
def support(message):
    # Создаем инлайн-кнопку для перехода к диалогу с пользователем @gorponk
    inline_keyboard = InlineKeyboardMarkup()
    inline_keyboard.add(InlineKeyboardButton("Написать в поддержку", url="https://t.me/gorponk"))

    # Отправляем сообщение с текстом и кнопкой "Написать в поддержку"
    bot.send_message(message.chat.id, "Написать в поддержку:", reply_markup=inline_keyboard)


# Функция для создания клавиатуры с меню
def create_menu_keyboard():
    keyboard = ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)  # Установка row_width=2
    buttons = ["Купить", "Полезная информация", "Мои заказы", "Поддержка"]
    for i in range(0, len(buttons), 2):  # Итерирование по кнопкам попарно
        button_row = buttons[i:i+2]  # Получение двух кнопок для каждого ряда
        keyboard.add(*[KeyboardButton(text=button) for button in button_row])
    return keyboard

# Запуск бота
bot.polling()
